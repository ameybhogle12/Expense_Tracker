import openpyxl
from openpyxl_image_loader import SheetImageLoader
import sys
import os

try:
    file_path = "C:/Amey/Projects/Flutter Projects/Expense_Tracker/test/Expense Tracker Testing Sheet.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Bugs"]
    
    # Try to load images
    image_loader = SheetImageLoader(sheet)
    
    image_count = 0
    # iterate over all cells that might have an image
    for row in range(1, 100):
        for col in range(1, 20):
            cell_name = f"{openpyxl.utils.get_column_letter(col)}{row}"
            if image_loader.image_in(cell_name):
                image = image_loader.get(cell_name)
                image_path = f"C:/Amey/Projects/Flutter Projects/Expense_Tracker/test/extracted_image_{image_count}.png"
                image.save(image_path)
                print(f"Extracted image to {image_path}")
                image_count += 1
                
    if image_count == 0:
        print("No images found in the Bugs sheet using this method.")
        
    # Also print all text from the sheet
    print("\n--- Text Content ---")
    for row in sheet.iter_rows(values_only=True):
        if any(row):
            print(row)
            
except Exception as e:
    print(f"Error: {e}")
